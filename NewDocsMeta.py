#!/usr/bin/python3
"""
Check the metadata for each ordinance and update if needed.
Since the metadata can cross-reference other Internet Archive
identifiers, the metadata can change as new items are uploaded.
"""

from openpyxl import load_workbook
from internetarchive import *
import IA_SQL
import argparse
import os
import glob
#import pickle
#import sqlite3
import tempfile
import shutil

parser = argparse.ArgumentParser()
parser.add_argument("coll_name", nargs='*', default=['1987'])
# args = parser.parse_args()
# input_name is list of strings
input_name = parser.parse_args().coll_name[0]

tmpDir = tempfile.mkdtemp(dir='/home/jghafa/archive/tmp',prefix='Ord-M-')+'/'

#Name of the Internet Archive collection target for uploads
CollectionName = 'citycouncilordinances'

# Title of the item in the collection.  This is the one people see.
Title = ''

#Unique indentifer for the upload, becomes the IA directory name
Identifier = ''

# Formatted ISO 8601, yyyy-mm-dd
Date = ''
Desc = ''
Notes= ''

# Fixed Internet Archive metadata fields
MediaType = 'texts'
Creator = 'City of Fort Wayne, Indiana'
License = 'http://creativecommons.org/licenses/by-nc-sa/4.0/'
Subject = ['Fort Wayne','Local Government','City Council']

def build_Bills_dict (Bills):
    """ Read Excel Ordinance data sheet and append it to a dictionary"""
    for ws in wb.worksheets:
        for row in ws.rows:
            # ignore rows where column B does not look like G-70-01
            if  row[1].value is not None and not row[1].data_type in 'b':
                if row[1].value[1] == '-' and row[1].value[4] == '-' :
                    bill_data = (row[1].value,row[2].value,row[3].value,row[6].value,row[13].value,row[14].value,row[15].value)
                    key = row[1].value.strip()
                    try:
                        print (Bills[key][0],'duplicate key')
                    except KeyError:
                        Bills[key] = bill_data
    return Bills

def Link(Title,URL,Display):
    """ return a <a> link """
    link='<a title="'+Title+'" href="'+URL+'" rel="nofollow">'+Display+'</a>'
    return link

BillType = {'A':'Appropriation','G':'General','R':'Resolution',
           'S':'Special','X':'Annexation','Z':'Zoning'}
brk = '<br />'

#wb = load_workbook(filename =
#    '//vs-videostorage/City Council Ordinances/Scanned Ordinance Index.xlsx')
wb = load_workbook(filename = '/media/smb/Scanned Ordinance Index.xlsx')
Bills = {}
Bills = build_Bills_dict (Bills)

# Read the Ordinance metadata from IA, starting the last ones entered.

SQLstring = '%'
if len(input_name) > 1:
    SQLstring = '%-' + input_name[-2:] + '-%'

#for c in reversed(CouncilOrdinance):
for row in IA_SQL.SearchItem('Ord',SQLstring):
    c = row[0]
    print(c+'     ',end='\r')
    bill = c.split('FWCityCouncil-Ordinance-')[1]

    final = Bills[bill][5].strftime("%Y-%m-%d")

    if Bills[bill][4] is None:
        intro = 'The Intro date not available'
    else:
        intro = Bills[bill][4].strftime("%Y-%m-%d")

    Desc = ('Bill: ' + bill + brk +
            'Type: ' + BillType[bill[0]] + brk +
            'Status: ' + Bills[bill][2] + brk +
            'Ordinance: ' + Bills[bill][1] + brk +
            Bills[bill][3] + brk +
            'Introduced: ' + intro + brk +
            'Final Disposition: ' + final)


    # Notes
    IntroID = 'FWCityCouncil-'+intro
    IntroLink =(Link(intro + ' Council Video',
        'https://archive.org/details/FWCityCouncil-'+intro,
        'Video of Council Introduction '+intro))

    #if IntroID in CouncilVideo:
    if IA_SQL.ItemExist('Video', IntroID):
        IntroLink += brk
    else:
        IntroLink = ''

    FinalID = 'FWCityCouncil-'+final
    FinalLink =(Link(final + ' Council Video',
            'https://archive.org/details/FWCityCouncil-'+final,
            'Video of Final Disposition '+final) + brk)

#    if FinalID in CouncilVideo:
    if IA_SQL.ItemExist('Video', FinalID):
        FinalLink += brk
    else:
        FinalLink = ''

    if Bills[bill][6] is None:
        SPDnotes = ''
    else:
        SPDnotes = Bills[bill][6]

    Notes = (IntroLink + FinalLink +
            'Document Notes:' + SPDnotes)

    # Get the metadata from IA
    item = get_item(c)
    update_meta = False

    try:
        IAdesc  = item.metadata['description']
    except KeyError:
        IAdesc  = ''
        print(c,'Desc  ***Missing Desc***')
    if Desc == IAdesc:
        pass
    else:
        print('Desc ',end='')
        update_meta = True

    try:
        IAnotes = item.metadata['notes']
    except KeyError:
        IAnotes = ''
        print(c,'Notes ***Missing Notes***')

    if Notes == IAnotes:
        pass
    else:
        print('Notes ',end='')
        update_meta = True


    if update_meta:
        r = item.modify_metadata(dict(description=Desc,notes=Notes))
        print (c,r,' IA metadata updated')

    # check title page of book, fix if needed
    item.download(files=c+'_scandata.xml',destdir=tmpDir,no_directory=True,retries=10)

    xml_In = open(tmpDir +          c +'_scandata.xml', 'r')
    xmlOut = open(tmpDir + 'new_' + c +'_scandata.xml', 'w')

    FirstPage=True
    modified=False
    for line in xml_In:
        rep_line = line
        if '<pageType>' in line:
            if FirstPage:
                FirstPage=False
                rep_line = line.replace('Normal','Title')
            else:
                rep_line = line.replace('Title','Normal')
        if not line == rep_line:
            modified = True
            line = rep_line
        xmlOut.write(line)

    xml_In.close()
    xmlOut.close()

    if modified:
        #upload xml file back
        os.remove(tmpDir +          c +'_scandata.xml')
        os.rename(tmpDir + 'new_' + c +'_scandata.xml',
                  tmpDir +          c +'_scandata.xml')
        r = item.upload(files=tmpDir+c+'_scandata.xml',retries=10)
        print (c,r,' XML updated')

    for tmpfile in glob.glob(tmpDir + '*.[xX][mM][lL]'):
        os.remove(tmpfile)

shutil.rmtree(tmpDir)
