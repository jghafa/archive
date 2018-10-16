#!/usr/bin/python3
"""
Download ordinances from AX
Write to the document server
Create a script file to upload files to AX
"""

from openpyxl import load_workbook
from internetarchive import *
import pickle
import glob
import argparse
from datetime import datetime
from time import strftime
import subprocess
#from pathlib import Path
import os


parser = argparse.ArgumentParser()
parser.add_argument("coll_name", nargs='*', default=['1970']) 
args = parser.parse_args()
# input_name is list of strings
input_name = args.coll_name

#Name of the Internet Archive collection target for uploads
TestIdSuffix = ''   #Set to '' when testing is done
#CollectionName = 'test_collection'
CollectionName = 'citycouncilordinances'

# Title of the item in the collection.  This is the one people see.
Title = ''

#Unique indentifer for the upload, becomes the IA directory name
Identifier = ''

# Formatted ISO 8601, yyyy-mm-dd
Date = ''
Desc = ''
Notes= ''

# Fixed Internet Archive metadata fields
MediaType = 'texts'
Creator = 'City of Fort Wayne, Indiana'
License = 'http://creativecommons.org/licenses/by-nc-sa/4.0/'
Subject = ['Fort Wayne','Local Government','City Council']



# Key to the Council Proceedings Index spreadsheet
# Col B, row[1],  Bills[file_name][0], Meeting Type
# Col C, row[2],  Bills[file_name][1], Date
# Col D, row[3],  Bills[file_name][2], Notes


# Key to the Scanned Ordinance Index spreadsheet
# Col B, row[1],  Bills[file_name][0], Bill
# Col C, row[2],  Bills[file_name][1], Ord
# Col D, row[3],  Bills[file_name][2], Status
# Col G, row[6],  Bills[file_name][3], Desc
# Col N, row[13], Bills[file_name][4], Intro
# Col O, row[14], Bills[file_name][5], Final
# Col P, row[15], Bills[file_name][6], Notes

def build_Bills_dict (Bills):
    """ Read Excel Ordinance data sheet and append it to a dictionary"""
    for ws in wb.worksheets:
        for row in ws.rows:
            # ignore rows where column B does not look like G-70-01
            if  row[1].value is not None and not row[1].data_type in 'b':
                if row[1].value[1] == '-' and row[1].value[4] == '-' :
                    bill_data = (row[1].value,row[2].value,row[3].value,row[6].value,row[13].value,row[14].value,row[15].value)
                    key = row[1].value.strip()
                    try:
                        print (Bills[key][0],'duplicate key')
                    except KeyError:
                        Bills[key] = bill_data
    return Bills

def build_Proceedings_dict (Proceedings, sheet):
    """ Read Excel Ordinance data sheet and append it to a dictionary"""
    Valid_Types = ['Council Proceeding','Other','Special']
    ws = wb[sheet]
    for row in ws.rows:
        # ignore rows where column B is not a valid meeting type
        if  row[1].value in Valid_Types:

            # Regular Meeting
            if row[1].value == 'Council Proceeding':
                key =('CR-' + str(row[2].value.day).zfill(2)+ '-'
                            + str(row[2].value.month).zfill(2) + '-'
                            + str(row[2].value.year))
            # Organzational Meeting    
            elif row[1].value == 'Other':
                key =('CO-' + str(row[2].value.day).zfill(2)+ '-'
                            + str(row[2].value.month).zfill(2) + '-'
                            + str(row[2].value.year))
            # Special Meeting
            elif row[1].value == 'Special':
                key =('CS-' + str(row[2].value.day).zfill(2)+ '-'
                            + str(row[2].value.month).zfill(2) + '-'
                            + str(row[2].value.year))
            Proceedings[key] = (row[3].value)
    return Proceedings

def Link(Title,URL,Display):
    """ return a <a> link """
    link='<a title="'+Title+'" target="_blank" href="'+URL+'">'+Display+'</a>'
    return link

def hyperlink(url,friendly):
    '''  create a hyperlink for a CSV that Excel will read'''
    h1 = '=hyperlink("'
    h2 = '","'
    h3 = '")'
    hyper = h1 + url + h2 + friendly + h3
    # remove the " so Excel import it
    # the Excel user will have to replace them back
    #wrap the link in quotes
    return '"'+hyper.replace('"','~')+'"'

BillType = {'A':'Appropriation','G':'General','R':'Resolution',
           'S':'Special','X':'Annexation','Z':'Zoning'}
brk = '<br>'

Bills = {}

picklefile = 'CouncilVideo.pickle'
try:
    CouncilVideo = pickle.load(open(picklefile, "rb"))
except (OSError, IOError) as e:
    print ('Reading councilmeeting collection')
    CouncilVideo = [item.metadata['identifier'] for item in search_items('collection:(councilmeetings)').iter_as_items()]
    pickle.dump(CouncilVideo, open(picklefile, "wb"), protocol=pickle.HIGHEST_PROTOCOL)

picklefile = 'CouncilOrdinance.pickle'
try:
    CouncilOrdinance = pickle.load(open(picklefile, "rb"))
except (OSError, IOError) as e:
    print ('Reading citycouncilordinance collection')
    CouncilOrdinance = [item.metadata['identifier'] for item in search_items('collection:(citycouncilordinances)').iter_as_items()]
    pickle.dump(CouncilOrdinance, open(picklefile, "wb"), protocol=pickle.HIGHEST_PROTOCOL)

# open log file
targetDir='/media/smb/PDFs/'+ input_name[0] + '/'
os.makedirs(targetDir, exist_ok=True)
log = open('../Documents/AXlog.txt', 'a')
AXlink = open(targetDir+'AXUpload-'+input_name[0]+'.txt', 'w')
log.write(datetime.now().strftime('%Y-%m-%d %H:%M:%S, ') + 'Start UpLoad \n')

#wb = load_workbook(filename =
#    '//vs-videostorage/City Council Ordinances/Council Proceedings Index.xlsx')
wb = load_workbook(filename = '/media/smb/Council Proceedings Index.xlsx')
Bills = build_Proceedings_dict (Bills, 'Council Proceedings')

#wb = load_workbook(filename =
#    '//vs-videostorage/City Council Ordinances/Scanned Ordinance Index.xlsx')
wb = load_workbook(filename = '/media/smb/Scanned Ordinance Index.xlsx')

Bills = build_Bills_dict (Bills)

# Read the file names
PATH = '/media/smb/Uploads'

# get all the files in and under PATH
files = [file for file in glob.glob(PATH + '/**/*.*', recursive=True)]

dirlist = []
fn_list = []
for fn in files:
    fn_list.append(fn.split('/')[-1])
    dirlist.append(fn.rstrip(fn.split('/')[-1]))
    
print ()
# loop thru file names
for f in range(len(fn_list)):
    file_name, file_ext = fn_list[f].split('.')
    if file_name == 'Thumbs':
        continue  # this a windows junk file
    if not file_ext.upper() == 'TIF':
        continue  # not a bill
    if 'Blueprint' in dirlist[f]:
        continue  # Skip the blueprints, they are batched with the primary
    prefix = file_name.split('-')[0]
    if prefix in ['CR','CS','CO']:
        continue # this is a council proceeding

    bill = file_name.split(' ')[0]
    Identifier = 'FWCityCouncil-Ordinance-'+bill+TestIdSuffix
    Title = 'Fort Wayne Ordinance '+bill

    try:
        if Bills[bill][6] is None:
            SPDnotes = ''
        else:
            SPDnotes = Bills[bill][6].replace('\n',' ')[0:253]

        #print (bill,Bills[bill][4],Bills[bill][5])
        final = Bills[bill][5].strftime("%Y-%m-%d")

        print('Identifier',Identifier,datetime.now().strftime('%Y-%m-%d %H:%M:%S'))

        # checking the year in the file path against a list years to be processed
        if dirlist[f].split('/')[4] in input_name or bill in input_name:
            FilePath = dirlist[f]+fn_list[f]
            #print('File Path',FilePath)
            #print(Bills[bill])

            item = get_item(Identifier)
            item.download(glob_pattern='*.pdf',destdir=targetDir,no_directory=True,retries=10)


            meta =(        Bills[bill][0]
                    +'|'  +Bills[bill][1]
                    +'|'  +Bills[bill][2]
                    +'|||'+Bills[bill][3].replace('\n',' ')[0:253]
                    +'|||||||'+Bills[bill][4].strftime("%m/%d/%Y")
                    +'|'+Bills[bill][5].strftime("%m/%d/%Y")
                    +'|'+SPDnotes+'|'
                    +'\n')
            #print(meta)
            AXlink.write(meta)
            AXlink.write('@@'+targetDir+bill+'.PDF'+'\n')

        
    except KeyError:
        print(dirlist[f], fn_list[f],'<<<<========== Not Found in spreadsheet')

log.close()
AXlink.close()
