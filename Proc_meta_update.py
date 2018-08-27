#!/usr/bin/python3
"""
Code to upload the council proceedings
"""

from openpyxl import load_workbook
from internetarchive import *
import os
import glob
import pickle
import tempfile
import shutil

tmpDir = tempfile.mkdtemp(dir='/home/jghafa/archive/tmp',prefix='Proc-M-')+'/'

# True for uploading files, false for debugging
update_IA = True

#
CollectionName = 'citycouncilproceedings'

# Title of the item in the collection.  This is the one people see.
Title = ''

#Unique indentifer for the upload, becomes the IA directory name
Identifier = ''

# Formatted ISO 8601, yyyy-mm-dd
Date = ''
Desc = ''
Notes= ''

# Fixed Internet Archive metadata fields
MediaType = 'texts'
Creator = 'City of Fort Wayne, Indiana'
License = 'http://creativecommons.org/licenses/by-nc-sa/4.0/'
Subject = ['Fort Wayne','Local Government','City Council']



def build_Proceedings_dict (Proceedings, sheet):
    """ Read Excel Ordinance data sheet and append it to a dictionary"""
    Valid_Types = ['Council Proceeding','Other','Special']
    ws = wb[sheet]
    for row in ws.rows:
        # ignore rows where column B is not a valid meeting type
        if  row[1].value in Valid_Types:

            # Regular Meeting
            if row[1].value == 'Council Proceeding':
                key =('CR-' + str(row[2].value.month).zfill(2) + '-'
                            + str(row[2].value.day).zfill(2)+ '-'
                            + str(row[2].value.year))
            # Organzational Meeting    
            elif row[1].value == 'Other':
                key =('CO-' + str(row[2].value.month).zfill(2) + '-'
                            + str(row[2].value.day).zfill(2)+ '-'
                            + str(row[2].value.year))
            # Special Meeting
            elif row[1].value == 'Special':
                key =('CS-' + str(row[2].value.month).zfill(2) + '-'
                            + str(row[2].value.day).zfill(2)+ '-'
                            + str(row[2].value.year))
            Proceedings[key] = (row[3].value)
    return Proceedings

def Link(Title,URL,Display):
    """ return a <a> link """
    link='<a title="'+Title+'" href="'+URL+'" rel="nofollow">'+Display+'</a>'
    return link

BillType = {'A':'Appropriation','G':'General','R':'Resolution',
           'S':'Special','X':'Annexation','Z':'Zoning'}
ProcType = {'CR':'Regular','CO':'Organizational','CS':'Special'}
brk = '<br>'

Procs = {}

picklefile = 'CouncilVideo.pickle'
try:
    CouncilVideo = pickle.load(open(picklefile, "rb"))
except (OSError, IOError) as e:
    print ('Reading councilmeeting collection')
    CouncilVideo = [item.metadata['identifier'] for item in search_items('collection:(councilmeetings)').iter_as_items()]
    pickle.dump(CouncilVideo, open(picklefile, "wb"), protocol=pickle.HIGHEST_PROTOCOL)

picklefile = 'CouncilOrdinance.pickle'
try:
    CouncilOrdinance = pickle.load(open(picklefile, "rb"))
except (OSError, IOError) as e:
    print ('Reading citycouncilordinance collection')
    CouncilOrdinance = [item.metadata['identifier'] for item in search_items('collection:(citycouncilordinances)').iter_as_items()]
    pickle.dump(CouncilOrdinance, open(picklefile, "wb"), protocol=pickle.HIGHEST_PROTOCOL)

picklefile = 'CouncilProceedings.pickle'
try:
    CouncilProceedings = pickle.load(open(picklefile, "rb"))
except (OSError, IOError) as e:
    print ('Reading citycouncilproceeding collection')
    CouncilProceedings = [item.metadata['identifier'] for item in search_items('collection:(citycouncilproceedings)').iter_as_items()]
    pickle.dump(CouncilProceedings, open(picklefile, "wb"), protocol=pickle.HIGHEST_PROTOCOL)



#wb = load_workbook(filename =
#    '//vs-videostorage/City Council Ordinances/Council Proceedings Index.xlsx')
wb = load_workbook(filename = '/media/smb/Council Proceedings Index.xlsx')
Procs = build_Proceedings_dict (Procs, 'Council Proceedings')

# Read the file names
PATH = '/media/smb/Uploads'


# Read the Ordinance metadata from IA
for c in CouncilProceedings:
    p_type = c.split('-')[2]
    p_yr   = c.split('-')[-3]
    p_mon  = c.split('-')[-2]
    p_day  = c.split('-')[-1]
    p_name = p_type + '-' + p_yr + '-' + p_mon + '-' + p_day
    spd_name = p_type + '-' + p_mon + '-' + p_day + '-' + p_yr

    if p_type == 'Index':
        continue

    Identifier = 'FWCityCouncil-Proceedings-'+p_name
    Title = 'Fort Wayne Council Proceedings '+p_name

    try:
        if Procs[spd_name] is None:
            SPDnotes = ''
        else:
            SPDnotes = Procs[spd_name]
    except KeyError:
        print(spd_name,'<<<<========== Not Found in spreadsheet')

    MeetDate = p_yr + '-' + p_mon + '-' + p_day
    MeetID = 'FWCityCouncil-'+ MeetDate
    
    if MeetID in CouncilVideo:
        MeetLink =(Link(MeetDate + ' Council Video',
            'https://archive.org/details/FWCityCouncil-'+MeetDate,
            'Video of Council Meeting '+MeetDate))
        MeetLink += brk
    else:
        MeetLink = ''

    Notes = (MeetLink + 'Notes: ' + SPDnotes)

    Desc = ProcType[p_type]+' Council Proceedings'

    Subject='Fort Wayne;'+ProcType[p_type]+' Council Proceedings'+';'+MeetDate

    # Get the metadata from IA
    item = get_item(c)
    update_meta = False

    try:
        IAdesc  = item.metadata['description']
    except KeyError:
        IAdesc  = ''
        print(c,'Desc  ***Missing***')
    if Desc == IAdesc:
        pass
    else:
        update_meta = True

    try:
        IAnotes = item.metadata['notes']
    except KeyError:
        IAnotes = ''
        print(c,'Notes ***Missing***')

    if Notes == IAnotes:
        pass
    else:
        # Update Notes
        update_meta = True

    if update_meta and update_IA:
        r = item.modify_metadata(dict(description=Desc,notes=Notes))
        print (r,' IA metadata updated')

    # check title page of book, fix if needed
    item.download(files=c+'_scandata.xml',destdir=tmpDir,no_directory=True,retries=10)

    xml_In = open(tmpDir +          c +'_scandata.xml', 'r')
    xmlOut = open(tmpDir + 'new_' + c +'_scandata.xml', 'w')

    FirstPage=True
    modified=False
    for line in xml_In:
        rep_line = line
        if '<pageType>' in line:
            if FirstPage:
                FirstPage=False
                rep_line = line.replace('Normal','Title')
            else:            
                rep_line = line.replace('Title','Normal')
        if not line == rep_line:
            modified = True
            line = rep_line
        xmlOut.write(line)
        
    xml_In.close()
    xmlOut.close()

    if modified:
        #upload xml file back
        os.remove(tmpDir +          c +'_scandata.xml')
        os.rename(tmpDir + 'new_' + c +'_scandata.xml',
                  tmpDir +          c +'_scandata.xml')
        if update_IA:
            r = item.upload(files=tmpDir+c+'_scandata.xml',retries=10)
            print (r,' XML updated')

    for tmpfile in glob.glob(tmpDir + '*.[xX][mM][lL]'):
        os.remove(tmpfile)

shutil.rmtree(tmpDir)
