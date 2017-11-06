#!/usr/bin/python3
"""
Update the video meta data
"""

from openpyxl import load_workbook
from internetarchive import *
import os
import glob
import pickle

tmpDir = '/home/jghafa/archive/tmp/'

#Name of the Internet Archive collection target for uploads
CollectionName = 'citycouncilordinances'

# Title of the item in the collection.  This is the one people see.
Title = ''

#Unique indentifer for the upload, becomes the IA directory name
Identifier = ''

# Formatted ISO 8601, yyyy-mm-dd
Date = ''
Desc = ''
Notes= ''

# Fixed Internet Archive metadata fields
MediaType = 'texts'
Creator = 'City of Fort Wayne, Indiana'
License = 'http://creativecommons.org/licenses/by-nc-sa/4.0/'
Subject = ['Fort Wayne','Local Government','City Council']

def build_Proceedings_dict (Proceedings, sheet):
    """ Read Excel Ordinance data sheet and append it to a dictionary"""
    Valid_Types = ['Council Proceeding','Other','Special']
    ws = wb[sheet]
    for row in ws.rows:
        # ignore rows where column B is not a valid meeting type
        if  row[1].value in Valid_Types:

            # Regular Meeting
            if row[1].value == 'Council Proceeding':
                key =('CR-' + str(row[2].value.month).zfill(2) + '-'
                            + str(row[2].value.day).zfill(2)+ '-'
                            + str(row[2].value.year))
            # Organzational Meeting    
            elif row[1].value == 'Other':
                key =('CO-' + str(row[2].value.month).zfill(2) + '-'
                            + str(row[2].value.day).zfill(2)+ '-'
                            + str(row[2].value.year))
            # Special Meeting
            elif row[1].value == 'Special':
                key =('CS-' + str(row[2].value.month).zfill(2) + '-'
                            + str(row[2].value.day).zfill(2)+ '-'
                            + str(row[2].value.year))
            Proceedings[key] = (row[3].value)
    return Proceedings

def build_Bills_dict (Bills):
    """ Read Excel Ordinance data sheet and append it to a dictionary"""
    for ws in wb.worksheets:
        for row in ws.rows:
            # ignore rows where column B does not look like G-70-01
            if  row[1].value is not None and not row[1].data_type in 'b':
                if row[1].value[1] == '-' and row[1].value[4] == '-' :
                    bill_data = (row[1].value,row[2].value,row[3].value,row[6].value,row[13].value,row[14].value,row[15].value)
                    key = row[1].value.strip()
                    try:
                        print (Bills[key][0],'duplicate key')
                    except KeyError:
                        Bills[key] = bill_data
    return Bills

def Link(Title,URL,Display):
    """ return a <a> link """
    link='<a title="'+Title+'" target="_blank" href="'+URL+'">'+Display+'</a>'
    return link

BillType = {'A':'Appropriation','G':'General','R':'Resolution',
           'S':'Special','X':'Annexation','Z':'Zoning'}
ProcType = {'CR':'Regular','CO':'Organizational','CS':'Special'}
brk = '<br>'

picklefile = 'CouncilVideo.pickle'
try:
    CouncilVideo = pickle.load(open(picklefile, "rb"))
except (OSError, IOError) as e:
    print ('Reading councilmeeting collection')
    CouncilVideo = [item.metadata['identifier'] for item in search_items('collection:(councilmeetings)').iter_as_items()]
    pickle.dump(CouncilVideo, open(picklefile, "wb"), protocol=pickle.HIGHEST_PROTOCOL)

picklefile = 'CouncilOrdinance.pickle'
try:
    CouncilOrdinance = pickle.load(open(picklefile, "rb"))
except (OSError, IOError) as e:
    print ('Reading citycouncilordinance collection')
    CouncilOrdinance = [item.metadata['identifier'] for item in search_items('collection:(citycouncilordinances)').iter_as_items()]
    pickle.dump(CouncilOrdinance, open(picklefile, "wb"), protocol=pickle.HIGHEST_PROTOCOL)

picklefile = 'CouncilProceedings.pickle'
try:
    CouncilProceedings = pickle.load(open(picklefile, "rb"))
except (OSError, IOError) as e:
    print ('Reading citycouncilproceeding collection')
    CouncilProceedings = [item.metadata['identifier'] for item in search_items('collection:(citycouncilproceedings)').iter_as_items()]
    pickle.dump(CouncilProceedings, open(picklefile, "wb"), protocol=pickle.HIGHEST_PROTOCOL)

wb = load_workbook(filename = '/media/smb/Council Proceedings Index.xlsx')
Procs = {}
Procs = build_Proceedings_dict (Procs, 'Council Proceedings')
wb = load_workbook(filename = '/media/smb/Scanned Ordinance Index.xlsx')
Bills = {}
Bills = build_Bills_dict (Bills)

Intro = {}
Final = {}
for p in Bills:
    #print (p, Bills[p][4], Bills[p][5])
    Introdate= (str(Bills[p][4].year)  +'-'+
                str(Bills[p][4].month).zfill(2) +'-'+
                str(Bills[p][4].day).zfill(2))
    Finaldate= (str(Bills[p][5].year)  +'-'+
                str(Bills[p][5].month).zfill(2) +'-'+
                str(Bills[p][5].day).zfill(2))
    # setdefault avoids the use of defaultdict
    Intro.setdefault(Introdate,[]).append(p)
    Final.setdefault(Finaldate,[]).append(p)

#x=input('p')

# loop through video
for v in CouncilVideo:
    #find the video date
    p_yr   = v.split('-')[-3]
    p_mon  = v.split('-')[-2]
    p_day  = v.split('-')[-1]
    p_date = p_mon + '-' + p_day + '-' + p_yr
    v_date = p_yr + '-' + p_mon + '-' + p_day
    o_date = p_yr[2:] + '-' + p_mon + '-' + p_day
    print (v)
    #x=input('next')
    #look for a matching ordinance
    Notes = ''
    if v_date in Intro:
        Notes += ('Introductions on '+v_date+brk)
        sortedlist = sorted(Intro[v_date])
        for ord_num in sortedlist:
            Notes += (Link(ord_num,
                       'https://archive.org/details/FWCityCouncil-Ordinance-'+ord_num,
                       ord_num + ' ' + BillType[ord_num.split('-')[0]])+brk)
    if v_date in Final:
        Notes += (brk+'Final Disposition on '+v_date+brk)
        sortedlist = sorted(Final[v_date])
        for ord_num in sortedlist:
            Notes += (Link(ord_num,
                       'https://archive.org/details/FWCityCouncil-Ordinance-'+ord_num,
                       ord_num + ' ' + BillType[ord_num.split('-')[0]])+brk)
    #look for a matching proceeding
    if 'CO-'+p_date in Procs:
        Notes += (brk+Link('Organizational Council Proceedings '+v_date,
                   'https://archive.org/details/FWCityCouncil-Proceedings-'+
                   'CO-'+v_date,
                   'Council Proceedings '+v_date+' ,Organizational')+brk)
    if 'CR-'+p_date in Procs:
        Notes += (brk+Link('Regular Council Proceedings '+v_date,
                   'https://archive.org/details/FWCityCouncil-Proceedings-'+
                   'CR-'+v_date,
                   'Council Proceedings '+v_date+', Regular')+brk)
    if 'CS-'+p_date in Procs:
        Notes += (brk+Link('Special Council Proceedings '+v_date,
                   'https://archive.org/details/FWCityCouncil-Proceedings-'+
                   'CS-'+v_date,
                   'Council Proceedings '+v_date+', Special')+brk)
    if Notes:
        #Identifier = 'FWCityCouncil-'+v_date
        item = get_item(v)

        #print(Notes)
        #print()
        try:
            IAnotes = item.metadata['notes']
        except KeyError:
            IAnotes = ''
            print(v,'Notes ***Missing***')

        if Notes == IAnotes:
            pass
        else:
            # Update Notes
            r = item.modify_metadata(dict(notes=Notes))
            print (v,r,' IA metadata updated')
