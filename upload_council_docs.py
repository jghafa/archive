#!/usr/bin/python3

from openpyxl import load_workbook
from internetarchive import *
import pickle
import glob
import argparse
from datetime import datetime
from time import strftime
import subprocess
import os
import tempfile

# True for uploading files, false for debugging
update_IA = True

parser = argparse.ArgumentParser()
parser.add_argument("coll_name", nargs='*', default=['1983']) 
args = parser.parse_args()
# input_name is list of strings
input_name = args.coll_name

#Name of the Internet Archive collection target for uploads
TestIdSuffix = ''   #Set to '' when testing is done
#CollectionName = 'test_collection'
CollectionName = 'citycouncilordinances'

# Title of the item in the collection.  This is the one people see.
Title = ''

#Unique indentifer for the upload, becomes the IA directory name
Identifier = ''

# Formatted ISO 8601, yyyy-mm-dd
Date = ''
Desc = ''
Notes= ''

# Fixed Internet Archive metadata fields
MediaType = 'texts'
Creator = 'City of Fort Wayne, Indiana'
License = 'http://creativecommons.org/licenses/by-nc-sa/4.0/'
Subject = ['Fort Wayne','Local Government','City Council']



# Key to the Council Proceedings Index spreadsheet
# Col B, row[1],  Bills[file_name][0], Meeting Type
# Col C, row[2],  Bills[file_name][1], Date
# Col D, row[3],  Bills[file_name][2], Notes


# Key to the Scanned Ordinance Index spreadsheet
# Col B, row[1],  Bills[file_name][0], Bill
# Col C, row[2],  Bills[file_name][1], Ord
# Col D, row[3],  Bills[file_name][2], Status
# Col G, row[6],  Bills[file_name][3], Desc
# Col N, row[13], Bills[file_name][4], Intro
# Col O, row[14], Bills[file_name][5], Final
# Col P, row[15], Bills[file_name][6], Notes

def build_Bills_dict (Bills):
    """ Read Excel Ordinance data sheet and append it to a dictionary"""
    for ws in wb.worksheets:
        for row in ws.rows:
            # ignore rows where column B does not look like G-70-01
            if  row[1].value is not None and not row[1].data_type in 'b':
                if row[1].value[1] == '-' and row[1].value[4] == '-' :
                    bill_data = (row[1].value,row[2].value,row[3].value,row[6].value,row[13].value,row[14].value,row[15].value)
                    key = row[1].value.strip()
                    try:
                        print (Bills[key][0],'duplicate key')
                    except KeyError:
                        Bills[key] = bill_data
                    if Bills[key][1] is None:
                        print (Bills[key][0],',Missing Ord Number')
                    if Bills[key][2] is None:
                        print (Bills[key][0],',Missing Bill Status')
                    if Bills[key][3] is None:
                        print (Bills[key][0],',Missing Bill Desc')
                    if Bills[key][4] is None:
                        print (Bills[key][0],',Missing Intro date')
                    if Bills[key][5] is None:
                        print (Bills[key][0],',Missing Final date')
    return Bills

def build_Proceedings_dict (Proceedings, sheet):
    """ Read Excel Ordinance data sheet and append it to a dictionary"""
    Valid_Types = ['Council Proceeding','Other','Special']
    ws = wb[sheet]
    for row in ws.rows:
        # ignore rows where column B is not a valid meeting type
        if  row[1].value in Valid_Types:

            # Regular Meeting
            if row[1].value == 'Council Proceeding':
                key =('CR-' + str(row[2].value.day).zfill(2)+ '-'
                            + str(row[2].value.month).zfill(2) + '-'
                            + str(row[2].value.year))
            # Organzational Meeting    
            elif row[1].value == 'Other':
                key =('CO-' + str(row[2].value.day).zfill(2)+ '-'
                            + str(row[2].value.month).zfill(2) + '-'
                            + str(row[2].value.year))
            # Special Meeting
            elif row[1].value == 'Special':
                key =('CS-' + str(row[2].value.day).zfill(2)+ '-'
                            + str(row[2].value.month).zfill(2) + '-'
                            + str(row[2].value.year))
            Proceedings[key] = (row[3].value)
    return Proceedings

def Link(Title,URL,Display):
    """ return a <a> link """
    link='<a title="'+Title+'" target="_blank" href="'+URL+'">'+Display+'</a>'
    return link

def hyperlink(url,friendly):
    '''  create a hyperlink for a CSV that Excel will read'''
    h1 = '=hyperlink("'
    h2 = '","'
    h3 = '")'
    hyper = h1 + url + h2 + friendly + h3
    # remove the " so Excel import it
    # the Excel user will have to replace them back
    #wrap the link in quotes
    return '"'+hyper.replace('"','~')+'"'

BillType = {'A':'Appropriation','G':'General','R':'Resolution',
           'S':'Special','X':'Annexation','Z':'Zoning'}
brk = '<br>'

Bills = {}

picklefile = 'CouncilVideo.pickle'
try:
    CouncilVideo = pickle.load(open(picklefile, "rb"))
except (OSError, IOError) as e:
    print ('Reading councilmeeting collection')
    CouncilVideo = [item.metadata['identifier'] for item in search_items('collection:(councilmeetings)').iter_as_items()]
    pickle.dump(CouncilVideo, open(picklefile, "wb"), protocol=pickle.HIGHEST_PROTOCOL)

picklefile = 'CouncilOrdinance.pickle'
try:
    CouncilOrdinance = pickle.load(open(picklefile, "rb"))
except (OSError, IOError) as e:
    print ('Reading citycouncilordinance collection')
    CouncilOrdinance = [item.metadata['identifier'] for item in search_items('collection:(citycouncilordinances)').iter_as_items()]
    pickle.dump(CouncilOrdinance, open(picklefile, "wb"), protocol=pickle.HIGHEST_PROTOCOL)


# open log file
log = open('../Documents/log.txt', 'a')
xlink = open('../Documents/Crosslink.txt', 'a')
log.write(datetime.now().strftime('%Y-%m-%d %H:%M:%S, ') + 'Start UpLoad \n')
xlink.write(datetime.now().strftime('%Y-%m-%d %H:%M:%S, ') + 'Start UpLoad \n')

xlink.write('Error,Bill,Intro,Intro Day,Final,Final Day,Notes' '\n')


#wb = load_workbook(filename =
#    '//vs-videostorage/City Council Ordinances/Council Proceedings Index.xlsx')
wb = load_workbook(filename = '/media/smb/Council Proceedings Index.xlsx')
Bills = build_Proceedings_dict (Bills, 'Council Proceedings')

#wb = load_workbook(filename =
#    '//vs-videostorage/City Council Ordinances/Scanned Ordinance Index.xlsx')
wb = load_workbook(filename = '/media/smb/Scanned Ordinance Index.xlsx')

Bills = build_Bills_dict (Bills)

# Read the file names
PATH = '/media/smb/Uploads'

tmpDir = tempfile.mkdtemp(dir='/home/jghafa/archive/tmp',prefix='Ord-U-')+'/'

# get all the files in and under PATH
files = [file for file in glob.glob(PATH + '/**/*.*', recursive=True)]

dirlist = []
fn_list = []
for fn in files:
    fn_list.append(fn.split('/')[-1])
    dirlist.append(fn.rstrip(fn.split('/')[-1]))
    
print ()
# loop thru file names
for f in range(len(fn_list)):
    file_name, file_ext = fn_list[f].split('.')
    if file_name == 'Thumbs':
        continue  # this a windows junk file
    if not file_ext.upper() == 'TIF':
        continue  # not a bill
    if 'Blueprint' in dirlist[f]:
        continue  # Skip the blueprints, they are batched with the primary
    prefix = file_name.split('-')[0]
    if prefix in ['CR','CS','CO']:
        continue # this is a council proceeding

    bill = file_name.split(' ')[0]
    #print(bill, end=' ')
    Identifier = 'FWCityCouncil-Ordinance-'+bill+TestIdSuffix
    Title = 'Fort Wayne Ordinance '+bill

    try:
        if Bills[bill][6] is None:
            SPDnotes = ''
        else:
            SPDnotes = Bills[bill][6]
        #print (bill,Bills[bill][4],Bills[bill][5])
        final = Bills[bill][5].strftime("%Y-%m-%d")

        if Bills[bill][4] is None:
            intro = 'The Intro date not available'
        else:
            intro = Bills[bill][4].strftime("%Y-%m-%d")

        Desc = ('Bill: ' + bill + brk +
                'Type: ' + BillType[bill[0]] + brk +
                'Status: ' + Bills[bill][2] + brk +
                'Ordinance: ' + Bills[bill][1] + brk +
                Bills[bill][3] + brk +
                'Introduced: ' + intro + brk +
                'Final Disposition: ' + final)
        
        IntroID = 'FWCityCouncil-'+intro
        IntroLink =(Link(intro + ' Council Video',
            'https://archive.org/details/FWCityCouncil-'+intro,
            'Video of Council Introduction '+intro))
        
        if IntroID in CouncilVideo:
            IntroLink += brk
        else:
            #print (bill)
            if Bills[bill][4].year in range(1981,2007):
                xlink.write('Missing Intro Video,'+
                  hyperlink(dirlist[f].replace('/media/smb/','\\\\vs-videostorage\\City Council Ordinances\\').replace('/','\\') + fn_list[f],bill)+ ',' +
                  hyperlink('https://archive.org/details/FWCityCouncil-'+intro, intro) + ',' +
                  '"=datevalue(indirect(address(row(),column()-1,4)))"' + ',' +                
                  hyperlink('https://archive.org/details/FWCityCouncil-'+final, final) + ',' +
                  '"=datevalue(indirect(address(row(),column()-1,4)))"' + ',' +                
                  '\n')
            IntroLink = ''
                        
        FinalID = 'FWCityCouncil-'+final
        FinalLink =(Link(final + ' Council Video',
                'https://archive.org/details/FWCityCouncil-'+final,
                'Video of Final Disposition '+final) + brk)

        if FinalID in CouncilVideo:
            FinalLink += brk
        else:
            #print (bill)
            if Bills[bill][5].year in range(1981,2007):
                xlink.write('Missing Final Video,'+
                  hyperlink(dirlist[f].replace('/media/smb/','\\\\vs-videostorage\\City Council Ordinances\\').replace('/','\\') + fn_list[f],bill)+ ',' +
                  hyperlink('https://archive.org/details/FWCityCouncil-'+intro, intro) + ',' +
                  '"=datevalue(indirect(address(row(),column()-1,4)))"' + ',' +                
                  hyperlink('https://archive.org/details/FWCityCouncil-'+final, final) + ',' +
                  '"=datevalue(indirect(address(row(),column()-1,4)))"' + ',' +                
                  '\n')
            FinalLink = ''

        Notes = (IntroLink + FinalLink +
                'Document Notes:' + SPDnotes)

        Subject='Fort Wayne;'+bill+';'+Bills[bill][1]

        if Identifier in CouncilOrdinance and not bill in input_name:
            print('Skipping',Identifier,datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
            continue

        print('Identifier',Identifier,datetime.now().strftime('%Y-%m-%d %H:%M:%S'))

        # checking the year in the file path against a list years to be processed
        listyear = dirlist[f].split('/')[4]
        if listyear in input_name or bill in input_name:
            FilePath = dirlist[f]+fn_list[f]
            print('File Path',FilePath)
            #print(brk)
            #print('title',Title)
            #print(brk)
            #print(MediaType,CollectionName,final,intro)
            #print(brk)
            #print(Desc)
            #print(brk)
            #print(Notes)
            #print(brk)
            #print(Subject)
                            
            md = dict(  collection = CollectionName, 
                        title      = Title,
                        mediatype  = MediaType, 
                        description= Desc,
                        creator    = Creator,
                        subject    = Subject,
                        licenseurl = License,
                        notes      = Notes,
                        date       = final)
            #print(md)

            convertList = glob.glob(dirlist[f] + bill + '*.[tT][iI][fF]')
            
            tifnum = 0
            for c in convertList:
                convertCmd = ('convert ' + c.replace(' ','\ ') + ' '
                              + final + '-' + str(tifnum).zfill(3) + '%03d.tif')
                #print(convertCmd)
                x = subprocess.run( [convertCmd],
                         cwd=tmpDir,
                         stdout=subprocess.DEVNULL,
                         shell=True)            
                tifnum += 1

            # Add the blueprints, if needed
            # should be bill instead of file_name
            if glob.glob('/media/smb/Uploads/Blueprints/'+bill+'*.[tT][iI][fF]'):
                convertCmd = ('convert ' + '/media/smb/Uploads/Blueprints/'
                              + bill +'*.[tT][iI][fF]'
                              +  ' ' + final + '-B%03d.tif' )
                #print(convertCmd)
                x = subprocess.run( [convertCmd],
                         cwd=tmpDir,
                         stdout=subprocess.DEVNULL,
                         shell=True)
                print('Blueprints Added to',Identifier)

            # Zip the TIFs into a single file to upload
            zipFile = tmpDir + Identifier + '_images.zip'
            zipCmd = 'zip ' + zipFile + ' *.[tT][iI][fF]'
            x = subprocess.run([zipCmd],
                     cwd=tmpDir,
                     stdout=subprocess.DEVNULL,
                     shell=True)

            # Send the files to IA
            if update_IA:
                try:
                    r = upload(Identifier, files=zipFile, metadata=md, 
                               retries=30, checksum=True) #retries_sleep=20,
                    print ('Status', r[0].status_code, zipFile)
                    log.write(datetime.now().strftime('%Y-%m-%d %H:%M:%S, ') + 
                              FilePath +' uploaded' + '\n')
                    CouncilOrdinance.append(Identifier) # Note to avoid further uploads
                    pickle.dump(CouncilOrdinance, open(picklefile, "wb"),
                                protocol=pickle.HIGHEST_PROTOCOL)

                except Exception as e:
                    print('Upload Failed on ', zipFile, e.message, e.args)
                    log.write(datetime.now().strftime('%Y-%m-%d %H:%M:%S, ') + 
                              FilePath +' failed' +  e.message + '\n')
                    continue
            else:
                z=input('update_IA is False')
                
            # Delete temp files')

            for tmpfile in glob.glob(tmpDir + '*.[tT][iI][fF]'):
                os.remove(tmpfile)
            for tmpfile in glob.glob(tmpDir + '*.[zZ][iI][pP]'):
                os.remove(tmpfile)

        
    except KeyError:
        print(dirlist[f], fn_list[f],'<<<<========== Not Found in spreadsheet')

log.close()
xlink.close()
