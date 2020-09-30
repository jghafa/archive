#!/usr/bin/python3
"""
Upload TFI mpg files to IA

"""

import glob
import os
import datetime
from internetarchive import *
from openpyxl import load_workbook
import tempfile
import shutil

tmpDir = tempfile.mkdtemp(dir='/home/jghafa/archive/tmp',prefix='TFI-')+'/'

PATH='/home/jghafa/archive/TFI/'
Break = '<br>'

#Name of the Internet Archive collection target for uploads
TestIdPrefix = ''   #Set to '' when testing is done
#CollectionName = 'test_collection'
CollectionName = 'theaterforideas'

#Read the spreadsheet for files to upload
wb = load_workbook(filename = PATH+'TFI.xlsx')
ws=wb[wb.sheetnames[0]] # read the leftmost worksheet
for row in ws.rows:
    # This metadata media type
    vidmediatype = 'movies'

    # Title of the item in the collection.  This is the one people see.
    vidTitle = ''

    #Unique indentifer for the upload, becomes the IA directory name
    Identifier = ''

    #Vidio date, used for the metadata, formated ISO 8601
    # yyyy-mm-dd
    vidDate = ''

    #video description, used for metadata, from spreadsheet via text file
    vidDesc = ''
    vidNotes= ''
    vidAudio=''

    #metadata for the producer
    vidCreator = ''
    vidDirector= ''

    #metadata for the Creative Commons license
    vidLicense = 'http://creativecommons.org/licenses/by-nc-sa/4.0/'

    #metadata for search fields
    vidSubject = ['Fort Wayne','Theater for Ideas','Public Access TV']

    try:
        FILENAME    ='Doran_'+str(int(row[0].value[0:2]))
    except:
        continue

    if os.path.isfile(PATH+FILENAME+'.mp4'):
        FILENAME += '.mp4'
    elif os.path.isfile(PATH+FILENAME+'.mpg'):
        FILENAME += '.mpg'
    else:
        print ('Missing ',PATH+FILENAME)
        continue
    
    DATE        =row[2].value
    DESCRIPTION =row[3].value
    TOPIC1      =row[4].value
    TOPIC2      =row[5].value
    TOPIC3      =row[6].value
    PARTICIPANT1=row[7].value
    PARTICIPANT2=row[8].value
    PARTICIPANT3=row[9].value
    PARTICIPANT4=row[10].value
    PARTICIPANT5=row[11].value
    PARTICIPANT6=row[12].value
    PRODUCER1   =row[13].value
    PRODUCER2   =row[14].value
    DIRECTOR    =row[15].value
    ASSTDIRECTOR=row[16].value
    FLOORMANAGER=row[17].value
    AUDIO       =row[18].value
    CAMERA1     =row[19].value
    CAMERA2     =row[20].value
    CAMERA3     =row[21].value
    NOTES       =row[22].value
    MUSIC       =row[23].value
    EDITOR      =row[24].value
    CREDITS     =row[25].value
    
    tempname = TestIdPrefix+'Theater for Ideas - '+row[1].value
    Identifier = tempname.replace('.mpg','').replace(',','-').replace(' - ','-').replace('- ','-').replace("'",'').replace(':','').replace('\xa0',' ').replace('  ',' ').replace(' ','-').replace('.mpg','')

    vidTitle = tempname.split('.')[0].split('   ')[0].strip()
    vidDesc = DESCRIPTION

#    print(tempname)
#    print(Identifier)
#    print(vidTitle)
#    x=input('name')
    if TOPIC1:
       vidSubject.append(TOPIC1) 
    if TOPIC2:
       vidSubject.append(TOPIC2) 
    if TOPIC3:
       vidSubject.append(TOPIC3) 

    if (   PARTICIPANT1 or PARTICIPANT2 or PARTICIPANT2
        or PARTICIPANT4 or PARTICIPANT5 or PARTICIPANT6):
        vidNotes += Break + Break + "Participants:" + Break
        if PARTICIPANT1:
            vidNotes += PARTICIPANT1 + Break
        if PARTICIPANT2:
            vidNotes += PARTICIPANT2 + Break
        if PARTICIPANT3:
            vidNotes += PARTICIPANT3 + Break
        if PARTICIPANT4:
            vidNotes += PARTICIPANT4 + Break
        if PARTICIPANT5:
            vidNotes += PARTICIPANT5 + Break
        if PARTICIPANT6:
            vidNotes += PARTICIPANT6 + Break
        vidNotes += Break
    if DIRECTOR:
        vidNotes +=  Break + DIRECTOR
    if ASSTDIRECTOR:
        vidNotes += Break + ASSTDIRECTOR
    if FLOORMANAGER:
        vidNotes += Break + FLOORMANAGER
    if CAMERA1:
        vidNotes += Break + "Cameras:" + CAMERA1
    if CAMERA2:
        vidNotes += Break + CAMERA2
    if CAMERA3:
        vidNotes += Break + CAMERA3
    if MUSIC:
        vidNotes += MUSIC + Break
    if NOTES:
        vidNotes = Break + NOTES

    if PRODUCER1:
        vidDirector = PRODUCER1
        vidCreator  = PRODUCER1
    if PRODUCER2 and PRODUCER1:
        vidDirector += Break
    if PRODUCER2:
        vidDirector += PRODUCER2

    vidAudio = AUDIO

    try:
        vidDate = DATE.strftime('%Y-%m-%d')
    except AttributeError:
        vidDate = DATE
        
    md = dict(  collection = CollectionName, 
                title      = vidTitle,
                mediatype  = vidmediatype, 
                description= vidDesc,
                creator    = vidCreator,
                director   = vidDirector,
                subject    = vidSubject,
                licenseurl = vidLicense,
                notes      = vidNotes,
                sound      = vidAudio,
                credits    = CREDITS,
                date       = vidDate
              )  
    #print(FILENAME   )
    #print(Identifier )
    #for m in md:
    #    print(m,md[m])
    #    print()
    #x=input('ready to upload')

    try:
        r = upload(Identifier, files=PATH+FILENAME, metadata=md, 
                   retries=30, checksum=True) #retries_sleep=20,
        print ('Status code', r[0].status_code, Identifier)
    except Exception as e:
        print ('Failed on ', Identifier, e.message, e.args)
        
shutil.rmtree(tmpDir)
