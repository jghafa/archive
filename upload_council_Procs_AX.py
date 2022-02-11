#!/usr/bin/python3
"""
Download the council proceedings from AX
Save to a spot on the document server
Create an AX import script to load the documents into AX
"""

from openpyxl import load_workbook
from internetarchive import *
import os
import glob
#import pickle
#import sqlite3
import IA_SQL
from datetime import datetime
from time import strftime
import subprocess
import argparse

parser = argparse.ArgumentParser()
parser.add_argument("coll_name", nargs='*', default=['1969']) 
args = parser.parse_args()
input_name = args.coll_name

tmpDir = '/home/jghafa/archive/tmp/'

#
CollectionName = 'citycouncilproceedings'

# Title of the item in the collection.  This is the one people see.
Title = ''

#Unique indentifer for the upload, becomes the IA directory name
Identifier = ''

# Formatted ISO 8601, yyyy-mm-dd
Date = ''
Desc = ''
Notes= ''

# Fixed Internet Archive metadata fields
MediaType = 'texts'
Creator = 'City of Fort Wayne, Indiana'
License = 'http://creativecommons.org/licenses/by-nc-sa/4.0/'
Subject = ['Fort Wayne','Local Government','City Council']



def build_Proceedings_dict (Proceedings, sheet):
    """ Read Excel Ordinance data sheet and append it to a dictionary"""
    Valid_Types = ['Council Proceeding','Other','Special']
    ws = wb[sheet]
    for row in ws.rows:
        # ignore rows where column B is not a valid meeting type
        if  row[1].value in Valid_Types:

            # Regular Meeting
            if row[1].value == 'Council Proceeding':
                key =('CR-' + str(row[2].value.month).zfill(2) + '-'
                            + str(row[2].value.day).zfill(2)+ '-'
                            + str(row[2].value.year))
            # Organzational Meeting    
            elif row[1].value == 'Other':
                key =('CO-' + str(row[2].value.month).zfill(2) + '-'
                            + str(row[2].value.day).zfill(2)+ '-'
                            + str(row[2].value.year))
            # Special Meeting
            elif row[1].value == 'Special':
                key =('CS-' + str(row[2].value.month).zfill(2) + '-'
                            + str(row[2].value.day).zfill(2)+ '-'
                            + str(row[2].value.year))
            if row[3].value is None:
                Proceedings[key] = ('')
            else:
                Proceedings[key] = (row[3].value)
    return Proceedings

def Link(Title,URL,Display):
    """ return a <a> link """
    link='<a title="'+Title+'" target="_blank" href="'+URL+'">'+Display+'</a>'
    return link

BillType = {'A':'Appropriation','G':'General','R':'Resolution',
           'S':'Special','X':'Annexation','Z':'Zoning'}
ProcType = {'CR':'Regular','CO':'Organizational','CS':'Special'}
brk = '<br>'

Procs = {}

# open log file
targetDir='/media/smb/PDFs/Proc'+ input_name[0] + '/'
os.makedirs(targetDir, exist_ok=True)
log = open('../Documents/AXlog.txt', 'a')
AXlink = open(targetDir+'AXUpload-Proc-'+input_name[0]+'.txt', 'w')
log.write(datetime.now().strftime('%Y-%m-%d %H:%M:%S, ') + 'Start UpLoad \n')

#wb = load_workbook(filename =
#    '//vs-videostorage/City Council Ordinances/Council Proceedings Index.xlsx')
wb = load_workbook(filename = '/media/smb/Council Proceedings Index.xlsx')
Procs = build_Proceedings_dict (Procs, 'Council Proceedings')

# Read the file names
PATH = '/media/smb/Uploads'

# Read the Ordinance metadata from IA, starting with recent uploads
#SQLstring = 'SELECT * FROM Proceeding WHERE locked = 0 ORDER BY item DESC'

# Read the Ordinance metadata from IA, new uploads first
#for c in reversed(CouncilProceedings):
#for row in SQL.execute(SQLstring):
for row in IA_SQL.SearchItem('Proceeding','%'):
    c = row[0]
    p_type = c.split('-')[2]
    p_yr   = c.split('-')[-3]
    p_mon  = c.split('-')[-2]
    p_day  = c.split('-')[-1]
    p_name = p_type + '-' + p_yr + '-' + p_mon + '-' + p_day
    spd_name = p_type + '-' + p_mon + '-' + p_day + '-' + p_yr

#    if not p_yr in input_name:
#        continue

    if p_yr in input_name or p_name in input_name:
        Identifier = 'FWCityCouncil-Proceedings-'+p_name
        # Get the PDF from IA
        item = get_item(c)
        item.download(glob_pattern='*.pdf',destdir=targetDir,no_directory=True,retries=10)

        meta =(       p_mon+'/'+p_day+'/'+p_yr
                +'|' +ProcType[p_type]
                +'|' +Procs[spd_name].replace('\n',' ')
                +'|' 
                +'\n')
        #print(meta)
        AXlink.write(meta)
        AXlink.write('@@'+targetDir+Identifier+'.PDF'+'\n')

log.close()
AXlink.close()
