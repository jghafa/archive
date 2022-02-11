#!/usr/bin/python3
"""
Code to upload the council proceedings
"""

from openpyxl import load_workbook
from internetarchive import *
#import pickle
#import sqlite3
import IA_SQL
import glob
import argparse
from datetime import datetime
from time import strftime
import subprocess
import os
import shutil
import tempfile

# True for uploading files, false for debugging
update_IA = True

parser = argparse.ArgumentParser()
parser.add_argument("coll_name", nargs='*', default=['1970']) 
args = parser.parse_args()
# input_name is list of strings
input_name = args.coll_name

#Name of the Internet Archive collection target for uploads
TestIdSuffix = ''   #Set to '' when testing is done
#CollectionName = 'test_collection'
CollectionName = 'citycouncilproceedings'

# Title of the item in the collection.  This is the one people see.
Title = ''

#Unique indentifer for the upload, becomes the IA directory name
Identifier = ''

# Formatted ISO 8601, yyyy-mm-dd
Date = ''
Desc = ''
Notes= ''

# Fixed Internet Archive metadata fields
MediaType = 'texts'
Creator = 'City of Fort Wayne, Indiana'
License = 'http://creativecommons.org/licenses/by-nc-sa/4.0/'
Subject = ['Fort Wayne','Local Government','City Council']

# Key to the Council Proceedings Index spreadsheet
# Col B, row[1],  Bills[file_name][0], Meeting Type
# Col C, row[2],  Bills[file_name][1], Date
# Col D, row[3],  Bills[file_name][2], Notes


# Key to the Scanned Ordinance Index spreadsheet
# Col B, row[1],  Bills[file_name][0], Bill
# Col C, row[2],  Bills[file_name][1], Ord
# Col D, row[3],  Bills[file_name][2], Status
# Col G, row[6],  Bills[file_name][3], Desc
# Col N, row[13], Bills[file_name][4], Intro
# Col O, row[14], Bills[file_name][5], Final
# Col P, row[15], Bills[file_name][6], Notes

def build_Bills_dict (Bills):
    """ Read Excel Ordinance data sheet and append it to a dictionary"""
    for ws in wb.worksheets:
        for row in ws.rows:
            # ignore rows where column B does not look like G-70-01
            if  row[1].value is not None and not row[1].data_type in 'b':
                if row[1].value[1] == '-' and row[1].value[4] == '-' :
                    bill_data = (row[1].value,row[2].value,row[3].value,row[6].value,row[13].value,row[14].value,row[15].value)
                    key = row[1].value.strip()
                    try:
                        print (Bills[key][0],'duplicate key')
                    except KeyError:
                        Bills[key] = bill_data
                    if Bills[key][1] is None:
                        print (Bills[key][0],',Missing Ord Number')
                    if Bills[key][2] is None:
                        print (Bills[key][0],',Missing Bill Status')
                    if Bills[key][3] is None:
                        print (Bills[key][0],',Missing Bill Desc')
                    if Bills[key][4] is None:
                        print (Bills[key][0],',Missing Intro date')
                    if Bills[key][5] is None:
                        print (Bills[key][0],',Missing Final date')
    return Bills

def build_Proceedings_dict (Proceedings, sheet):
    """ Read Excel Ordinance data sheet and append it to a dictionary"""
    Valid_Types = ['Council Proceeding','Other','Special']
    ws = wb[sheet]
    for row in ws.rows:
        # ignore rows where column B is not a valid meeting type
        if  row[1].value in Valid_Types:

            # Regular Meeting
            if row[1].value == 'Council Proceeding':
                key =('CR-' + str(row[2].value.month).zfill(2) + '-'
                            + str(row[2].value.day).zfill(2)+ '-'
                            + str(row[2].value.year))
            # Organzational Meeting    
            elif row[1].value == 'Other':
                key =('CO-' + str(row[2].value.month).zfill(2) + '-'
                            + str(row[2].value.day).zfill(2)+ '-'
                            + str(row[2].value.year))
            # Special Meeting
            elif row[1].value == 'Special':
                key =('CS-' + str(row[2].value.month).zfill(2) + '-'
                            + str(row[2].value.day).zfill(2)+ '-'
                            + str(row[2].value.year))
            Proceedings[key] = (row[3].value)
    return Proceedings

def Link(Title,URL,Display):
    """ return a <a> link """
    link='<a title="'+Title+'" target="_blank" href="'+URL+'">'+Display+'</a>'
    return link

def hyperlink(url,friendly):
    '''  create a hyperlink for a CSV that Excel will read'''
    h1 = '=hyperlink("'
    h2 = '","'
    h3 = '")'
    hyper = h1 + url + h2 + friendly + h3
    # remove the " so Excel import it
    # the Excel user will have to replace them back
    #wrap the link in quotes
    return '"'+hyper.replace('"','~')+'"'

BillType = {'A':'Appropriation','G':'General','R':'Resolution',
           'S':'Special','X':'Annexation','Z':'Zoning'}
ProcType = {'CR':'Regular','CO':'Organizational','CS':'Special'}
brk = '<br />'

Procs = {}

"""
SQLconn = sqlite3.connect('Council.sqlite')
SQL = SQLconn.cursor()

Lock=True
Unlock=False
def LockItem(itemtype, bill, locked):
    ''' update the locked status of the item'''
    insstring = 'INSERT OR REPLACE into Ordinance values (?,?)'
    if itemtype[0] == 'P':
        insstring = 'INSERT OR REPLACE into Proceeding values (?,?)'
    if itemtype[0] == 'V':
        insstring = 'INSERT OR REPLACE into Video values (?,?)'
    SQL.execute(insstring,(bill,locked) )
    SQLconn.commit()

def RemoveItem(itemtype, bill):
    ''' Remove from SQL if upload failed '''
    selstring = 'DELETE FROM Ordinance WHERE item = (?);'
    if itemtype[0] == 'P':
        selstring = 'DELETE FROM Proceeding WHERE item = (?);'
    if itemtype[0] == 'V':
        selstring = 'DELETE FROM Video WHERE item = (?);'
    SQL.execute(selstring,(bill,) )
    SQLconn.commit()

def ItemExist(itemtype, bill):
    ''' Return True if the item exists, False if not '''
    selstring = 'SELECT * FROM Ordinance WHERE item = (?);'
    if itemtype[0] == 'P':
        selstring = 'SELECT * FROM Proceeding WHERE item = (?);'
    if itemtype[0] == 'V':
        selstring = 'SELECT * FROM Video WHERE item = (?);'
    for row in SQL.execute(selstring, (bill,) ):
        return True
    return False
"""


"""
picklefile = 'CouncilVideo.pickle'
try:
    CouncilVideo = pickle.load(open(picklefile, "rb"))
except (OSError, IOError) as e:
    print ('Reading councilmeeting collection')
    CouncilVideo = [item.metadata['identifier'] for item in search_items('collection:(councilmeetings)').iter_as_items()]
    pickle.dump(CouncilVideo, open(picklefile, "wb"), protocol=pickle.HIGHEST_PROTOCOL)

picklefile = 'CouncilOrdinance.pickle'
try:
    CouncilOrdinance = pickle.load(open(picklefile, "rb"))
except (OSError, IOError) as e:
    print ('Reading citycouncilordinance collection')
    CouncilOrdinance = [item.metadata['identifier'] for item in search_items('collection:(citycouncilordinances)').iter_as_items()]
    pickle.dump(CouncilOrdinance, open(picklefile, "wb"), protocol=pickle.HIGHEST_PROTOCOL)

picklefile = 'CouncilProceedings.pickle'
try:
    CouncilProceedings = pickle.load(open(picklefile, "rb"))
except (OSError, IOError) as e:
    print ('Reading citycouncilproceeding collection')
    CouncilProceedings = [item.metadata['identifier'] for item in search_items('collection:(citycouncilproceedings)').iter_as_items()]
    pickle.dump(CouncilProceedings, open(picklefile, "wb"), protocol=pickle.HIGHEST_PROTOCOL)
"""

# open log file
log = open('../Documents/log.txt', 'a')
#xlink = open('../Documents/Crosslink.txt', 'a')
log.write(datetime.now().strftime('%Y-%m-%d %H:%M:%S, ') + 'Start UpLoad \n')
#xlink.write(datetime.now().strftime('%Y-%m-%d %H:%M:%S, ') + 'Start UpLoad \n')

#xlink.write('Error,Bill,Intro,Intro Day,Final,Final Day,Notes' '\n')


#wb = load_workbook(filename =
#    '//vs-videostorage/City Council Ordinances/Council Proceedings Index.xlsx')
wb = load_workbook(filename = '/media/smb/Council Proceedings Index.xlsx')
Procs = build_Proceedings_dict (Procs, 'Council Proceedings')

#wb = load_workbook(filename =
#    '//vs-videostorage/City Council Ordinances/Scanned Ordinance Index.xlsx')
#wb = load_workbook(filename = '/media/smb/Scanned Ordinance Index.xlsx')

#Bills = build_Bills_dict (Bills)

# Read the file names
PATH = '/media/smb/Uploads'

# get all the files in and under PATH
files = [file for file in glob.glob(PATH + '/**/*.*', recursive=True)]

dirlist = []
fn_list = []
for fn in files:
    fn_list.append(fn.split('/')[-1])
    dirlist.append(fn.rstrip(fn.split('/')[-1]))

tmpDir = tempfile.mkdtemp(dir='/home/jghafa/archive/tmp',prefix='Proc-U-')+'/'
    
print ()
# loop thru file names
for f in range(len(fn_list)):
    file_name, file_ext = fn_list[f].split('.')
    if file_name == 'Thumbs':
        continue  # this a windows junk file
    if not file_ext.upper() == 'TIF':
        continue  # not a bill
    if 'Blueprint' in dirlist[f]:
        continue  # Skip the blueprints, they are batched with the primary
    proc_name = file_name.split(' ')[0]
    p_type = proc_name.split('-')[0]
    if not p_type in ['CR','CS','CO']:
        continue # this is a council proceeding
    p_mon  = proc_name.split('-')[1]
    p_day  = proc_name.split('-')[2]
    p_yr   = proc_name.split('-')[3]

    p_name = p_type + '-' + p_yr + '-' + p_mon + '-' + p_day
    Identifier = 'FWCityCouncil-Proceedings-'+p_name+TestIdSuffix
    Title = 'Fort Wayne Council Proceedings '+p_name
    #print (Procs[file_name])

    try:
        if Procs[proc_name] is None:
            SPDnotes = ''
        else:
            SPDnotes = Procs[proc_name]

        MeetDate = p_yr + '-' + p_mon + '-' + p_day
        MeetID = 'FWCityCouncil-'+ MeetDate
        MeetLink =(Link(MeetDate + ' Council Video',
            'https://archive.org/details/FWCityCouncil-'+MeetDate,
            'Video of Council Introduction '+MeetDate))
        
        #if MeetID in CouncilVideo:
        if IA_SQL.ItemExist('V','FWCityCouncil-'+MeetID):
            MeetLink += brk
        else:
            MeetLink = ''

        Notes = (MeetLink + 'Notes: ' + SPDnotes)

        Desc = ProcType[p_type]+' Council Proceedings'

        Subject='Fort Wayne;'+ProcType[p_type]+' Council Proceedings'+';'+MeetDate

        #if Identifier in CouncilProceedings and not p_name in input_name:
        if IA_SQL.ItemExist('Proc', Identifier) and not p_name in input_name:
            #print('Skipping',Identifier,datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
            continue

        IA_SQL.LockItem('Proc', Identifier, IA_SQL.Lock)
        print('Identifier',Identifier,datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
        # checking the year in the file path against a list years to be processed
        if dirlist[f].split('/')[4][:4] in input_name or p_name in input_name:
            FilePath = dirlist[f]+fn_list[f]
            print(dirlist[f].split('/')[4][:4],input_name ,p_name )
            print('File Path',FilePath)
                            
            md = dict(  collection = CollectionName, 
                        title      = Title,
                        mediatype  = MediaType, 
                        description= Desc,
                        creator    = Creator,
                        subject    = Subject,
                        licenseurl = License,
                        notes      = Notes,
                        date       = MeetDate)
            #print(md)

            convertList = glob.glob(dirlist[f] + proc_name + '*.[tT][iI][fF]')
            
            tifnum = 0
            for c in convertList:
                # c.replace escapes spaces in the file name
                convertCmd = ('convert ' + c.replace(' ','\ ') + ' '
                              + p_name + '-' + str(tifnum) + '%03d.tif')
                #print(convertCmd)
                x = subprocess.run( [convertCmd],
                         cwd=tmpDir,
                         stdout=subprocess.DEVNULL,
                         shell=True)            
                tifnum += 1

            # Add the blueprints, if needed
            # should be proc_name instead of file_name
            
            if glob.glob('/media/smb/Uploads/Blueprints/'+proc_name+'*.[tT][iI][fF]'):
                convertCmd = ('convert ' + '/media/smb/Uploads/Blueprints/'
                              + proc_name +'*.[tT][iI][fF]'
                              +  ' ' + p_name + '-B%03d.tif' )
                #print(convertCmd)
                x = subprocess.run( [convertCmd],
                         cwd=tmpDir,
                         stdout=subprocess.DEVNULL,
                         shell=True)
                print('Blueprints Added to',Identifier)

            # Zip the TIFs into a single file to upload
            zipFile = tmpDir + Identifier + '_images.zip'
            zipCmd = 'zip ' + zipFile + ' *.[tT][iI][fF]'
            #print (zipCmd)
            x = subprocess.run([zipCmd],
                     cwd=tmpDir,
                     stdout=subprocess.DEVNULL,
                     shell=True)

            # Send the files to IA
            if update_IA:
                try:
                    r = upload(Identifier, files=zipFile, metadata=md, 
                               retries=30, checksum=True) #retries_sleep=20,
                    print (r[0].status_code, zipFile)
                    log.write(datetime.now().strftime('%Y-%m-%d %H:%M:%S, ') + 
                              FilePath +' uploaded' + '\n')
                    IA_SQL.LockItem('Proc', Identifier, IA_SQL.Unlock)
                    #picklefile = 'CouncilProceedings.pickle'
                    #CouncilProceedings.append(Identifier) # Note to avoid further uploads
                    #pickle.dump(CouncilProceedings, open(picklefile, "wb"),
                    #            protocol=pickle.HIGHEST_PROTOCOL)

                except Exception as e:
                    print('Upload Failed on ', zipFile, e.message, e.args)
                    log.write(datetime.now().strftime('%Y-%m-%d %H:%M:%S, ') + 
                              FilePath +' failed' +  e.message + '\n')
                    IA_SQL.RemoveItem('Proc', Identifier)
                    continue
            else:
                z=input('update_IA is False')
                
            # Delete temp files')

            for tmpfile in glob.glob(tmpDir + '*.[tT][iI][fF]'):
                os.remove(tmpfile)
            for tmpfile in glob.glob(tmpDir + '*.[zZ][iI][pP]'):
                os.remove(tmpfile)

        
    except KeyError:
        print(dirlist[f], fn_list[f],'<<<<========== Not Found in spreadsheet')

log.close()
#xlink.close()
shutil.rmtree(tmpDir)
